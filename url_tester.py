"""
URL Tester Application - Optimized Version
A tool for testing multiple URLs and reporting any non-200 status codes.
Supports two modes: 1) Defined URL list 2) Sitemap parsing
Optimized to reduce executable size by removing pandas dependency.
"""

import requests
import time
import sys
import os
from datetime import datetime
from pathlib import Path
import logging
import xml.etree.ElementTree as ET
from urllib.parse import urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# Set UTF-8 encoding for Windows console
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

class URLTester:
    def __init__(self, mode="defined", input_file=None, output_file=None):
        """Initialize the URL Tester"""
        self.mode = mode  # "defined" or "sitemap"
        
        # Set input file based on mode
        if input_file is None:
            self.input_file = "urls_to_test.xlsx" if mode == "defined" else "sitemaps.xlsx"
        else:
            self.input_file = input_file
        
        # Generate output filename with timestamp if not provided
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_file = f"test_results_{timestamp}.xlsx"
        else:
            self.output_file = output_file
        
        self.root_url = ""
        self.urls = []
        self.errors = []
        
        # Setup logging (console only, no file)
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def load_urls(self):
        """Load URLs based on mode (defined list or sitemap)"""
        if self.mode == "defined":
            return self._load_defined_urls()
        else:  # sitemap mode
            return self._load_sitemap_urls()
    
    def _load_defined_urls(self):
        """Load URLs from the defined URL list Excel file using openpyxl"""
        try:
            if not Path(self.input_file).exists():
                self.logger.error(f"Input file '{self.input_file}' not found!")
                print(f"\nERROR: File '{self.input_file}' not found!")
                print("Please make sure the file exists in the same folder as this application.")
                input("\nPress Enter to exit...")
                sys.exit(1)
            
            # Load workbook with openpyxl
            wb = load_workbook(self.input_file, read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            if 'root' not in headers or 'url' not in headers:
                self.logger.error("Excel file must have 'root' and 'url' columns!")
                print("\nERROR: Excel file must have 'root' and 'url' columns!")
                input("\nPress Enter to exit...")
                sys.exit(1)
            
            root_idx = headers.index('root')
            url_idx = headers.index('url')
            
            # Read data rows
            root_found = False
            urls_list = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not root_found and row[root_idx]:
                    self.root_url = str(row[root_idx]).strip()
                    root_found = True
                
                if row[url_idx]:
                    urls_list.append(str(row[url_idx]).strip())
            
            self.urls = urls_list
            wb.close()
            
            self.logger.info(f"Loaded {len(self.urls)} URLs to test")
            self.logger.info(f"Root URL: {self.root_url}")
            
            print(f"\n[OK] Successfully loaded {len(self.urls)} URLs from defined list")
            print(f"[OK] Root URL: {self.root_url}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading URLs: {str(e)}")
            print(f"\nERROR loading file: {str(e)}")
            input("\nPress Enter to exit...")
            sys.exit(1)
    
    def _load_sitemap_urls(self):
        """Load sitemap URLs and parse them to get actual URLs"""
        try:
            if not Path(self.input_file).exists():
                self.logger.error(f"Input file '{self.input_file}' not found!")
                print(f"\nERROR: File '{self.input_file}' not found!")
                print("Please make sure the file exists in the same folder as this application.")
                input("\nPress Enter to exit...")
                sys.exit(1)
            
            # Load workbook with openpyxl
            wb = load_workbook(self.input_file, read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            if 'sitemap_url' not in headers:
                self.logger.error("Excel file must have 'sitemap_url' column!")
                print("\nERROR: Excel file must have 'sitemap_url' column!")
                input("\nPress Enter to exit...")
                sys.exit(1)
            
            sitemap_idx = headers.index('sitemap_url')
            
            # Read sitemap URLs
            sitemap_urls = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[sitemap_idx]:
                    sitemap_urls.append(str(row[sitemap_idx]).strip())
            
            wb.close()
            
            print(f"\n[OK] Found {len(sitemap_urls)} sitemap(s) to parse")
            print("[INFO] Fetching URLs from sitemaps...")
            
            all_urls = []
            for idx, sitemap_url in enumerate(sitemap_urls, 1):
                print(f"[INFO] Parsing sitemap {idx}/{len(sitemap_urls)}: {sitemap_url}")
                urls = self._parse_sitemap(sitemap_url)
                all_urls.extend(urls)
                print(f"[OK] Found {len(urls)} URLs in this sitemap")
            
            # Remove duplicates
            self.urls = list(set(all_urls))
            
            self.logger.info(f"Loaded {len(self.urls)} unique URLs from sitemaps")
            
            print(f"\n[OK] Total unique URLs to test: {len(self.urls)}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading sitemaps: {str(e)}")
            print(f"\nERROR loading sitemaps: {str(e)}")
            input("\nPress Enter to exit...")
            sys.exit(1)
    
    def _parse_sitemap(self, sitemap_url):
        """Parse a sitemap XML and extract URLs (non-recursive, only direct URLs)"""
        try:
            response = requests.get(sitemap_url, timeout=30)
            response.raise_for_status()
            
            urls = []
            root = ET.fromstring(response.content)
            
            # Handle namespace
            namespaces = {
                'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9',
                'news': 'http://www.google.com/schemas/sitemap-news/0.9',
                'image': 'http://www.google.com/schemas/sitemap-image/1.1'
            }
            
            # Check if this is a sitemap index (contains other sitemaps)
            sitemaps = root.findall('ns:sitemap', namespaces)
            if sitemaps:
                # This is a sitemap index - skip it, don't crawl deeper
                print(f"[INFO] Skipped sitemap index (contains other sitemaps): {sitemap_url}")
                print(f"[INFO] If you need URLs from this, add the specific sitemap URLs to sitemaps.xlsx")
                return []
            else:
                # This is a regular sitemap with URLs - extract them
                url_elements = root.findall('ns:url', namespaces)
                for url_element in url_elements:
                    loc = url_element.find('ns:loc', namespaces)
                    if loc is not None and loc.text:
                        urls.append(loc.text)
            
            return urls
            
        except Exception as e:
            self.logger.warning(f"Error parsing sitemap {sitemap_url}: {str(e)}")
            print(f"[WARNING] Could not parse sitemap {sitemap_url}: {str(e)}")
            return []
    
    def test_url(self, url):
        """Test a single URL and return status code (without delay, for concurrent use)"""
        try:
            # Combine root URL with relative URL if needed
            if self.root_url and not url.startswith(('http://', 'https://')):
                full_url = self.root_url.rstrip('/') + '/' + url.lstrip('/')
            else:
                full_url = url
            
            # Send request with timeout
            response = requests.get(
                full_url, 
                timeout=10,
                allow_redirects=True,
                headers={'User-Agent': 'URL-Tester/1.0'}
            )
            
            status_code = response.status_code
            
            return status_code, full_url, None
            
        except requests.exceptions.Timeout:
            return 'TIMEOUT', full_url, 'Request timed out'
        except requests.exceptions.ConnectionError:
            return 'CONNECTION_ERROR', full_url, 'Connection failed'
        except requests.exceptions.TooManyRedirects:
            return 'TOO_MANY_REDIRECTS', full_url, 'Too many redirects'
        except Exception as e:
            return 'ERROR', full_url, str(e)
    
    def test_all_urls(self, delay=0.1, max_workers=20):
        """Test all URLs concurrently - threads continuously pick up work as they finish"""
        print(f"\n[INFO] Starting URL tests...")
        print(f"[INFO] Max concurrent requests: {max_workers}")
        print(f"[INFO] Target rate: ~{int(1/delay)} requests/second")
        print("=" * 60)
        
        total = len(self.urls)
        success_count = 0
        error_count = 0
        completed = 0
        start_time = time.time()
        
        # Thread-safe counters
        lock = threading.Lock()
        
        def update_progress():
            nonlocal completed
            with lock:
                completed += 1
                if completed % 10 == 0 or completed == total:
                    progress = (completed / total) * 100
                    elapsed = time.time() - start_time
                    rate = completed / elapsed if elapsed > 0 else 0
                    print(f"Progress: {completed}/{total} ({progress:.1f}%) - Success: {success_count} | Errors: {error_count} | Rate: {rate:.1f} req/s")
        
        # Use ThreadPoolExecutor - threads will continuously pull work
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            
            # Submit all tasks - executor manages them efficiently
            for idx, url in enumerate(self.urls):
                future = executor.submit(self.test_url, url)
                futures[future] = (url, idx + 1)
                
                # Optional: Control submission rate to avoid overwhelming the system
                # Remove this if you want maximum speed
                if delay > 0 and idx < total - 1:
                    time.sleep(delay)
            
            # Process as they complete (not in submission order!)
            for future in as_completed(futures):
                url, idx = futures[future]
                try:
                    status_code, full_url, error_msg = future.result()
                    
                    # If not 200, log it as an error
                    if status_code != 200:
                        with lock:
                            error_count += 1
                        self.errors.append({
                            'original_url': url,
                            'full_url': full_url,
                            'status_code': status_code,
                            'error_message': error_msg if error_msg else '',
                            'tested_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        })
                        self.logger.warning(f"[{idx}/{total}] {status_code} - {full_url}")
                    else:
                        with lock:
                            success_count += 1
                    
                    update_progress()
                    
                except Exception as e:
                    with lock:
                        error_count += 1
                    self.logger.error(f"Error processing URL {url}: {str(e)}")
                    update_progress()
        
        elapsed = time.time() - start_time
        print("=" * 60)
        print(f"\n[OK] Testing complete!")
        print(f"  Total URLs tested: {total}")
        print(f"  Successful (200): {success_count}")
        print(f"  Errors: {error_count}")
        print(f"  Total time: {elapsed:.1f} seconds")
        print(f"  Average rate: {total/elapsed:.1f} requests/second")
        
        self.logger.info(f"Testing complete - Success: {success_count}, Errors: {error_count}")
    
    def save_results(self):
        """Save error results to Excel file using openpyxl"""
        if not self.errors:
            print("\n[SUCCESS] No errors found! All URLs returned status 200.")
            self.logger.info("No errors to report")
            return
        
        try:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Errors"
            
            # Write headers
            headers = ['original_url', 'full_url', 'status_code', 'error_message', 'tested_at']
            ws.append(headers)
            
            # Write data rows
            for error in self.errors:
                row = [
                    error['original_url'],
                    error['full_url'],
                    str(error['status_code']),
                    error['error_message'],
                    error['tested_at']
                ]
                ws.append(row)
            
            # Auto-adjust column widths
            for idx, col in enumerate(headers, 1):
                column_letter = get_column_letter(idx)
                max_length = len(col)
                
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(self.output_file)
            
            print(f"\n[OK] Error report saved to: {self.output_file}")
            self.logger.info(f"Results saved to {self.output_file}")
            
        except Exception as e:
            self.logger.error(f"Error saving results: {str(e)}")
            print(f"\nERROR saving results: {str(e)}")
    
    def run(self, delay=0.1, max_workers=10):
        """Main execution method"""
        print("=" * 60)
        print("           URL TESTER APPLICATION")
        print("=" * 60)
        
        # Load URLs from file
        self.load_urls()
        
        # Test all URLs
        self.test_all_urls(delay, max_workers)
        
        # Save results
        self.save_results()
        
        print("\n" + "=" * 60)
        print("Testing completed successfully!")
        print("=" * 60)


def main():
    """Main entry point"""
    try:
        print("=" * 60)
        print("           URL TESTER APPLICATION")
        print("=" * 60)
        print("\nSelect testing mode:")
        print("  1. Defined URL list (from urls_to_test.xlsx)")
        print("  2. Sitemap parsing (from sitemaps.xlsx)")
        print()
        
        # Get user choice
        while True:
            choice = input("Enter your choice (1 or 2): ").strip()
            if choice in ['1', '2']:
                break
            print("Invalid choice. Please enter 1 or 2.")
        
        # Set mode based on choice
        mode = "defined" if choice == "1" else "sitemap"
        
        print()
        print("=" * 60)
        
        # Create and run tester
        tester = URLTester(mode=mode)
        # 100ms delay between request starts, max 20 concurrent requests
        # Set delay=0 for maximum speed (no rate limiting)
        tester.run(delay=0.1, max_workers=20)
        
    except KeyboardInterrupt:
        print("\n\n[WARNING] Testing interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\nERROR: Unexpected error: {str(e)}")
        logging.error(f"Unexpected error: {str(e)}", exc_info=True)
        input("\nPress Enter to exit...")
        sys.exit(1)
    
    # Wait for user before closing
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()

