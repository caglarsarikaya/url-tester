"""Excel file reading and writing utilities"""

from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class ExcelReader:
    """Handles reading data from Excel files"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
    
    def exists(self) -> bool:
        """Check if file exists"""
        return self.file_path.exists()
    
    def read_rows(self, required_columns: List[str]) -> List[Dict[str, str]]:
        """
        Read rows from Excel file as dictionaries
        
        Args:
            required_columns: List of required column names
            
        Returns:
            List of dictionaries with column names as keys
            
        Raises:
            ValueError: If required columns are missing
        """
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Validate required columns
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            wb.close()
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Get column indices
        col_indices = {col: headers.index(col) for col in headers if col}
        
        # Read data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for col_name, col_idx in col_indices.items():
                value = row[col_idx]
                row_dict[col_name] = str(value).strip() if value else None
            rows.append(row_dict)
        
        wb.close()
        return rows


class ExcelWriter:
    """Handles writing data to Excel files"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
    
    def write_data(self, headers: List[str], rows: List[Dict[str, str]]):
        """
        Write data to Excel file with auto-adjusted column widths
        
        Args:
            headers: List of column headers
            rows: List of dictionaries containing row data
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Errors"
        
        # Write headers
        ws.append(headers)
        
        # Write data rows
        for row_dict in rows:
            row = [row_dict.get(header, '') for header in headers]
            ws.append(row)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, headers)
        
        # Save workbook
        wb.save(self.file_path)
    
    def _adjust_column_widths(self, ws, headers: List[str]):
        """Adjust column widths based on content"""
        for idx, col in enumerate(headers, 1):
            column_letter = get_column_letter(idx)
            max_length = len(col)
            
            for cell in ws[column_letter]:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set width with max limit
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

